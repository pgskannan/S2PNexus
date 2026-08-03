"""Unit tests: FS grade bands + Excel generate/import round-trip + tamper."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.services.excel_registration import (
    QUESTIONNAIRE_VERSION,
    SHEET_INSTRUCTIONS,
    SHEET_SUPPLIER_INFO,
    TEMPLATE_VERSION,
    _module_sheet_name,
    generate_registration_workbook,
    parse_and_validate_workbook,
)
from app.services.template_engine import (
    FS_REGISTRATION_GRADE_BANDS,
    grade_for_score,
)


class TestFsGradeBands:
    def test_boundaries(self):
        assert grade_for_score(Decimal("90"), bands=FS_REGISTRATION_GRADE_BANDS) == "A"
        assert grade_for_score(Decimal("89.99"), bands=FS_REGISTRATION_GRADE_BANDS) == "B"
        assert grade_for_score(Decimal("75"), bands=FS_REGISTRATION_GRADE_BANDS) == "B"
        assert grade_for_score(Decimal("74.99"), bands=FS_REGISTRATION_GRADE_BANDS) == "C"
        assert grade_for_score(Decimal("50"), bands=FS_REGISTRATION_GRADE_BANDS) == "C"
        assert grade_for_score(Decimal("49.99"), bands=FS_REGISTRATION_GRADE_BANDS) == "D"
        assert grade_for_score(Decimal("0"), bands=FS_REGISTRATION_GRADE_BANDS) == "D"

    def test_module_dispatch(self):
        assert grade_for_score(Decimal("75"), module="supplier_registration_core") == "B"
        assert grade_for_score(Decimal("75"), module="supplier_request") == "C"
        assert grade_for_score(Decimal("59"), module="supplier_request") == "F"


def _fake_template(module_code: str, questions: list[dict]):
    qs = [
        SimpleNamespace(
            question_key=q["key"],
            question_type=q.get("type", "text"),
            question_text=q.get("text", q["key"]),
            options=q.get("options"),
            mandatory_flag=q.get("mandatory", True),
            visible_flag=True,
            scoring_rule=q.get("scoring"),
            visibility_rule=None,
        )
        for q in questions
    ]
    return SimpleNamespace(
        module=f"supplier_registration_{module_code}",
        sections=[SimpleNamespace(visibility_rule=None, questions=qs)],
    )


def _fake_registration(**kwargs):
    defaults = dict(
        id=uuid4(),
        supplier_id=uuid4(),
        company_name="Acme LLC",
        legal_name="Acme LLC",
        address_line1="1 Main",
        address_line2=None,
        city="Austin",
        state_province="TX",
        postal_code="78701",
        country="US",
        tax_id="12-3456789",
        bank_account_number="123456789",
        bank_routing_number="021000021",
        primary_contact_name="Pat",
        primary_contact_email="pat@acme.example",
    )
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def _expected_sheets(module_codes: list[str]) -> list[str]:
    return [SHEET_INSTRUCTIONS, SHEET_SUPPLIER_INFO] + [_module_sheet_name(c) for c in module_codes]


class TestExcelRoundTrip:
    def test_generate_reimport_clean(self):
        reg = _fake_registration()
        templates = {
            "core": _fake_template(
                "core",
                [
                    {
                        "key": "years_in_business",
                        "type": "numeric",
                        "text": "Years",
                        "mandatory": True,
                        "scoring": {"weight": 10, "threshold": 1, "above": 10, "below": 0},
                    },
                    {
                        "key": "publicly_traded",
                        "type": "yes_no",
                        "text": "Public?",
                        "options": ["yes", "no"],
                        "mandatory": True,
                    },
                ],
            )
        }
        data, structure_hash = generate_registration_workbook(reg, "STD_VENDOR", templates)
        assert structure_hash
        assert len(data) > 100

        wb = load_workbook(BytesIO(data))
        ws = wb[_module_sheet_name("core")]
        ws["D2"] = 10
        ws["D3"] = "yes"
        out = BytesIO()
        wb.save(out)

        result = parse_and_validate_workbook(
            out.getvalue(),
            reg,
            expected_hash=structure_hash,
            expected_template_version=TEMPLATE_VERSION,
            expected_questionnaire_version=QUESTIONNAIRE_VERSION,
            expected_sheets=_expected_sheets(["core"]),
            templates_by_module_code=templates,
        )
        assert result.ok, [(f.category, f.rule, f.actual) for f in result.failures]
        assert result.answers_by_module["core"]["years_in_business"] == 10
        assert result.answers_by_module["core"]["publicly_traded"] == "yes"

    def test_tamper_locked_cell(self):
        reg = _fake_registration()
        templates = {"core": _fake_template("core", [{"key": "q1", "text": "Q1", "mandatory": False}])}
        data, structure_hash = generate_registration_workbook(reg, "STD_VENDOR", templates)
        wb = load_workbook(BytesIO(data))
        wb[SHEET_SUPPLIER_INFO]["A2"] = "TAMPERED"
        buf = BytesIO()
        wb.save(buf)
        result = parse_and_validate_workbook(
            buf.getvalue(),
            reg,
            expected_hash=structure_hash,
            expected_template_version=TEMPLATE_VERSION,
            expected_questionnaire_version=QUESTIONNAIRE_VERSION,
            expected_sheets=_expected_sheets(["core"]),
            templates_by_module_code=templates,
        )
        assert not result.ok
        cats = {f.category for f in result.failures}
        assert "tampered_locked_cell" in cats or "structural" in cats

    def test_tamper_remove_column(self):
        reg = _fake_registration()
        templates = {"core": _fake_template("core", [{"key": "q1", "text": "Q1", "mandatory": False}])}
        data, structure_hash = generate_registration_workbook(reg, "STD_VENDOR", templates)
        wb = load_workbook(BytesIO(data))
        wb[SHEET_SUPPLIER_INFO].delete_cols(3)
        buf = BytesIO()
        wb.save(buf)
        result = parse_and_validate_workbook(
            buf.getvalue(),
            reg,
            expected_hash=structure_hash,
            expected_template_version=TEMPLATE_VERSION,
            expected_questionnaire_version=QUESTIONNAIRE_VERSION,
            expected_sheets=_expected_sheets(["core"]),
            templates_by_module_code=templates,
        )
        assert not result.ok
        assert any(f.category in {"structural", "version_mismatch"} for f in result.failures)

    def test_bad_email_and_country(self):
        reg = _fake_registration()
        templates = {"core": _fake_template("core", [{"key": "q1", "text": "Q1", "mandatory": False}])}
        data, structure_hash = generate_registration_workbook(reg, "STD_VENDOR", templates)
        wb = load_workbook(BytesIO(data))
        ws = wb[SHEET_SUPPLIER_INFO]
        ws["C2"] = "Acme"
        ws["E2"] = "ZZ"
        ws["J2"] = "not-an-email"
        buf = BytesIO()
        wb.save(buf)
        result = parse_and_validate_workbook(
            buf.getvalue(),
            reg,
            expected_hash=structure_hash,
            expected_template_version=TEMPLATE_VERSION,
            expected_questionnaire_version=QUESTIONNAIRE_VERSION,
            expected_sheets=_expected_sheets(["core"]),
            templates_by_module_code=templates,
        )
        assert not result.ok
        rules = {f.rule for f in result.failures}
        # Rule names come from the importer — assert categories too.
        cats = {f.category for f in result.failures}
        assert "invalid_format" in cats or "email_format" in rules or any("email" in (f.rule or "") for f in result.failures)
        assert "invalid_format" in cats or any("country" in (f.rule or "") for f in result.failures)
