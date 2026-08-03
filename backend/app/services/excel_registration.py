"""Excel-based Supplier Registration: template generation + import engine (FS 13-16).

Foundation decision (already made, restated here for anyone reading this file
in isolation): questionnaire content comes from the Template Framework
(models/template.py, services/template_engine.py) via the
supplier_registration_* modules and MODULE_CODE_TO_TEMPLATE
(models/supplier_type.py) -- this module does not define a second
questionnaire grammar. It only knows how to lay TemplateQuestion rows out as
Excel cells and read them back.

Generation (FS 13.1/14) -- generate_registration_workbook():
    Sheet 1 "Instructions": versions + how-to-fill legend.
    Sheet 2 "Supplier Information": columns A-K, SupplierID/SupplierType/
        TemplateVersion locked, everything else editable.
    Sheet 3+: one sheet per questionnaire module, named after the module's
        short code Title-cased (e.g. "core" -> "Core"). Columns A-H;
        QuestionID/ModuleID hidden, Response/Comments editable, everything
        else locked. ScoreFormula is a fixed placeholder string, never the
        real weight (exposing weights in a workbook a supplier can open is
        an integrity leak, not just a UX one).

Structural integrity (FS 15.1) -- the "hash signature":
    A structural payload (sheet names, headers, hidden-column state, and the
    text of every locked cell) is extracted from the *actual* openpyxl
    Workbook object via _extract_sheet_meta(), then hashed with SHA-256.
    Both generation and import call the same extractor, so the import side
    is comparing "what this exact file's structure canonically looks like"
    against "the hash recorded at send time" -- it never has to re-derive an
    expectation from live templates that may have moved on since send.

Import (FS 16.2) -- parse_and_validate_workbook():
    Structural validation runs to completion FIRST (sheet set, column
    headers, hidden columns, the three ground-truth locked cells, version
    fields, and the structure hash). Any structural failure short-circuits
    the import before a single Response cell is read -- a tampered file must
    never be partially trusted. Only a structurally clean file proceeds to
    FS 15.2 field validation (mandatory, dropdown membership, email, numeric,
    bank regex, ISO country).

Known, documented simplification: openpyxl sheet/cell protection
(`ws.protection.sheet = True`) is a UX guard only -- any spreadsheet
application can trivially remove it, with or without a password, which this
batch doesn't set. The actual security boundary is server-side: the
structure hash and the locked-cell ground-truth comparisons below, not the
Excel protection flags themselves.
"""

from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from typing import TYPE_CHECKING, Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation

if TYPE_CHECKING:
    from app.models.supplier_registration import SupplierRegistration
    from app.models.template import TemplateDefinition

# ---------------------------------------------------------------------------
# Versions (Phase 4 stub constants; Phase 2 wiring reads these as defaults)
# ---------------------------------------------------------------------------

TEMPLATE_VERSION = "1.0"
QUESTIONNAIRE_VERSION = "1.0"

# ---------------------------------------------------------------------------
# Sheet / column layout (FS Section 14)
# ---------------------------------------------------------------------------

SHEET_INSTRUCTIONS = "Instructions"
SHEET_SUPPLIER_INFO = "Supplier Information"

SUPPLIER_INFO_COLUMNS = (
    "SupplierID",
    "SupplierType",
    "LegalName",
    "Address",
    "Country",
    "TaxID",
    "BankAccountNumber",
    "BankRoutingNumber",
    "ContactName",
    "ContactEmail",
    "TemplateVersion",
)
SUPPLIER_INFO_LOCKED = frozenset({"SupplierID", "SupplierType", "TemplateVersion"})
# FS Section 14 doesn't mark individual Supplier Information fields
# mandatory beyond the locked three; this is this batch's own reasonable
# default for what a usable registration record needs, documented rather
# than silently assumed.
SUPPLIER_INFO_MANDATORY = frozenset({"LegalName", "Country", "ContactEmail"})

QUESTIONNAIRE_COLUMNS = (
    "QuestionID",
    "ModuleID",
    "QuestionText",
    "Response",
    "AllowedValues",
    "MandatoryFlag",
    "ScoreFormula",
    "Comments",
)
QUESTIONNAIRE_HIDDEN_COLUMNS = ("A", "B")
QUESTIONNAIRE_EDITABLE_COLUMNS = (4, 8)  # Response, Comments
SCORE_FORMULA_PLACEHOLDER = "LOCKED"

# ---------------------------------------------------------------------------
# Field validation defaults (FS Section 15.2)
# ---------------------------------------------------------------------------

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
# Defaults per the implementation prompt; not every bank format matches a US
# ABA routing number, but this is the documented default for this batch.
BANK_ACCOUNT_REGEX = re.compile(r"^[A-Za-z0-9]{6,34}$")
BANK_ROUTING_REGEX = re.compile(r"^\d{9}$")  # US ABA routing number

# Practical ISO 3166-1 alpha-2 subset (~55 common trading-partner codes)
# rather than the full ~250-entry list -- enough to catch typos/garbage
# without this module owning a full country registry.
ISO_COUNTRY_CODES = frozenset(
    {
        "US", "CA", "MX", "GB", "IE", "FR", "DE", "ES", "PT", "IT",
        "NL", "BE", "LU", "CH", "AT", "SE", "NO", "DK", "FI", "PL",
        "CZ", "SK", "HU", "RO", "BG", "GR", "HR", "SI", "EE", "LV",
        "LT", "IN", "CN", "JP", "KR", "SG", "HK", "TW", "TH", "VN",
        "MY", "ID", "PH", "AU", "NZ", "BR", "AR", "CL", "CO", "ZA",
        "EG", "AE", "SA", "IL", "TR", "RU", "UA",
    }
)


# ---------------------------------------------------------------------------
# Result types (FS Section 16.4/16.5)
# ---------------------------------------------------------------------------


@dataclass
class ValidationFailure:
    """One row of FS 16.4's error taxonomy, structural or field-level."""

    category: str  # structural | version_mismatch | mandatory_missing | invalid_dropdown | invalid_format | tampered_locked_cell | unknown_question | unknown_module
    sheet: Optional[str]
    cell: Optional[str]
    rule: str
    expected: Optional[str]
    actual: Optional[str]


@dataclass
class ImportResult:
    """Outcome of parse_and_validate_workbook (FS 16.5's three outputs)."""

    ok: bool
    failures: list[ValidationFailure] = field(default_factory=list)
    supplier_info: dict[str, Any] = field(default_factory=dict)
    answers_by_module: dict[str, dict[str, Any]] = field(default_factory=dict)
    error_report_bytes: Optional[bytes] = None
    import_summary_text: Optional[str] = None


# ---------------------------------------------------------------------------
# Structure hash (FS Section 15.1)
# ---------------------------------------------------------------------------


def canonical_structure_payload(sheet_meta: dict[str, Any]) -> dict[str, Any]:
    """Normalize a raw sheet-metadata dict into the exact shape that gets hashed.

    Round-tripping through json.dumps(..., sort_keys=True) makes key order
    (and therefore the resulting hash) independent of dict insertion order.
    Both the generator and the importer build sheet_meta by walking a live
    openpyxl Workbook via the same _extract_sheet_meta(), so the *values*
    already agree when nothing was tampered with -- this function only pins
    down *serialization*.
    """
    return json.loads(json.dumps(sheet_meta, sort_keys=True, default=str))


def compute_structure_hash(payload: dict[str, Any]) -> str:
    """SHA-256 hex digest of the canonical structural payload."""
    encoded = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _module_sheet_name(module_code: str) -> str:
    """Short code -> sheet name, e.g. "core" -> "Core" (FS 13.1 Sheet 3+)."""
    return module_code.title()


def _module_code_from_sheet(sheet_name: str) -> str:
    """Inverse of _module_sheet_name -- exact for our single-word codes."""
    return sheet_name.lower()


def _extract_sheet_meta(wb: Workbook) -> dict[str, Any]:
    """Walk a live Workbook and pull out everything considered "structural".

    Used by both generate_registration_workbook (on the just-built in-memory
    workbook) and parse_and_validate_workbook (on the loaded-from-bytes
    workbook), so both sides compute the hash the exact same way.
    """
    meta: dict[str, Any] = {"sheets": list(wb.sheetnames)}

    if SHEET_INSTRUCTIONS in wb.sheetnames:
        ws = wb[SHEET_INSTRUCTIONS]
        meta["instructions"] = {
            "template_version": ws["B2"].value,
            "questionnaire_version": ws["B3"].value,
        }

    if SHEET_SUPPLIER_INFO in wb.sheetnames:
        ws = wb[SHEET_SUPPLIER_INFO]
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(SUPPLIER_INFO_COLUMNS) + 1)]
        meta["supplier_information"] = {
            "headers": headers,
            "locked_values": {
                header: ws.cell(row=2, column=idx + 1).value
                for idx, header in enumerate(SUPPLIER_INFO_COLUMNS)
                if header in SUPPLIER_INFO_LOCKED
            },
        }

    modules: dict[str, Any] = {}
    for name in wb.sheetnames:
        if name in (SHEET_INSTRUCTIONS, SHEET_SUPPLIER_INFO):
            continue
        ws = wb[name]
        module_code = _module_code_from_sheet(name)
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(QUESTIONNAIRE_COLUMNS) + 1)]
        hidden = [c for c in QUESTIONNAIRE_HIDDEN_COLUMNS if ws.column_dimensions[c].hidden]
        questions: list[dict[str, Any]] = []
        row = 2
        while ws.cell(row=row, column=1).value not in (None, ""):
            questions.append(
                {
                    "question_id": ws.cell(row=row, column=1).value,
                    "module_id": ws.cell(row=row, column=2).value,
                    "question_text": ws.cell(row=row, column=3).value,
                    "mandatory_flag": ws.cell(row=row, column=6).value,
                    "score_formula": ws.cell(row=row, column=7).value,
                }
            )
            row += 1
        modules[module_code] = {
            "sheet_name": name,
            "headers": headers,
            "hidden_columns": hidden,
            "questions": questions,
        }
    meta["modules"] = modules
    return meta


# ---------------------------------------------------------------------------
# Generation (FS Section 13.1 / 14)
# ---------------------------------------------------------------------------


def generate_registration_workbook(
    registration: "SupplierRegistration",
    supplier_type_code: str,
    templates_by_module_code: dict[str, "TemplateDefinition"],
    *,
    template_version: str = TEMPLATE_VERSION,
    questionnaire_version: str = QUESTIONNAIRE_VERSION,
) -> tuple[bytes, str]:
    """Build the locked workbook sent to a supplier.

    `templates_by_module_code` maps short module codes (e.g. "core", "tax")
    to their resolved TemplateDefinition (already picked via
    get_effective_template() by the caller) -- one sheet is generated per
    entry, in dict order.

    Returns (xlsx_bytes, structure_hash). structure_hash is computed from
    the workbook that was actually just built, not re-derived separately,
    so it is guaranteed to match what parse_and_validate_workbook computes
    from an untouched round-trip of the same file.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default "Sheet"

    _build_instructions_sheet(wb, template_version, questionnaire_version, templates_by_module_code)
    _build_supplier_info_sheet(wb, registration, supplier_type_code, template_version)
    for module_code, template in templates_by_module_code.items():
        _build_module_sheet(wb, module_code, template)

    sheet_meta = _extract_sheet_meta(wb)
    structure_hash = compute_structure_hash(canonical_structure_payload(sheet_meta))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), structure_hash


def _build_instructions_sheet(
    wb: Workbook,
    template_version: str,
    questionnaire_version: str,
    templates_by_module_code: dict[str, "TemplateDefinition"],
) -> None:
    ws = wb.create_sheet(SHEET_INSTRUCTIONS)
    ws["A1"] = "S2PNexus Supplier Registration"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "TemplateVersion"
    ws["B2"] = template_version
    ws["A3"] = "QuestionnaireVersion"
    ws["B3"] = questionnaire_version

    lines = [
        "How to fill this workbook:",
        "1. Complete the 'Supplier Information' sheet -- editable fields only.",
        "2. Complete each questionnaire module sheet's 'Response' and 'Comments' columns.",
        "3. Fields where MandatoryFlag = 'Yes' must not be left blank.",
        "4. Do not unhide, reorder, rename, or delete any sheet or column.",
        "5. Do not edit locked (grey/protected) cells -- doing so fails import validation.",
        "",
        "Mandatory legend: MandatoryFlag = 'Yes' means the Response column is required.",
        "",
        "Questionnaire modules included in this workbook:",
    ]
    row = 5
    for line in lines:
        ws.cell(row=row, column=1, value=line)
        row += 1
    for module_code in templates_by_module_code:
        ws.cell(row=row, column=1, value=f"- {_module_sheet_name(module_code)}")
        row += 1

    ws.column_dimensions["A"].width = 95
    ws.protection.sheet = True


def _build_supplier_info_sheet(
    wb: Workbook,
    registration: "SupplierRegistration",
    supplier_type_code: str,
    template_version: str,
) -> None:
    ws = wb.create_sheet(SHEET_SUPPLIER_INFO)
    for idx, header in enumerate(SUPPLIER_INFO_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    address = getattr(registration, "address_line1", None) or ""
    address_line2 = getattr(registration, "address_line2", None)
    if address_line2:
        address = f"{address}, {address_line2}" if address else address_line2
    city = getattr(registration, "city", None)
    if city:
        address = f"{address}, {city}" if address else city

    values = {
        "SupplierID": str(registration.id),
        "SupplierType": supplier_type_code,
        "LegalName": getattr(registration, "legal_name", None) or getattr(registration, "company_name", None) or "",
        "Address": address,
        "Country": getattr(registration, "country", None) or "",
        "TaxID": getattr(registration, "tax_id", None) or "",
        "BankAccountNumber": getattr(registration, "bank_account_number", None) or "",
        "BankRoutingNumber": getattr(registration, "bank_routing_number", None) or "",
        "ContactName": getattr(registration, "primary_contact_name", None) or "",
        "ContactEmail": getattr(registration, "primary_contact_email", None) or "",
        "TemplateVersion": template_version,
    }
    for idx, header in enumerate(SUPPLIER_INFO_COLUMNS, start=1):
        cell = ws.cell(row=2, column=idx, value=values[header])
        cell.protection = Protection(locked=header in SUPPLIER_INFO_LOCKED)

    country_dv = DataValidation(
        type="list",
        formula1='"' + ",".join(sorted(ISO_COUNTRY_CODES)) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid country",
        error="Select a valid ISO 3166-1 alpha-2 country code.",
    )
    ws.add_data_validation(country_dv)
    country_dv.add("E2:E2")

    ws.protection.sheet = True


def _build_module_sheet(wb: Workbook, module_code: str, template: "TemplateDefinition") -> None:
    ws = wb.create_sheet(_module_sheet_name(module_code))
    for idx, header in enumerate(QUESTIONNAIRE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)

    questions = [q for section in template.sections for q in section.questions]
    row = 2
    for question in questions:
        options = getattr(question, "options", None) or (
            ["Yes", "No"] if question.question_type == "yes_no" else []
        )
        allowed_text = ", ".join(str(o) for o in options)

        ws.cell(row=row, column=1, value=question.question_key)
        ws.cell(row=row, column=2, value=module_code)
        ws.cell(row=row, column=3, value=question.question_text)
        ws.cell(row=row, column=4, value=getattr(question, "default_value", None) or "")
        ws.cell(row=row, column=5, value=allowed_text)
        ws.cell(row=row, column=6, value="Yes" if question.mandatory_flag else "No")
        # Integrity, not just UX: the real scoring weight never appears in a
        # workbook a supplier can open. See module docstring.
        ws.cell(row=row, column=7, value=SCORE_FORMULA_PLACEHOLDER)
        ws.cell(row=row, column=8, value="")

        if options:
            formula = '"' + ",".join(str(o) for o in options) + '"'
            if len(formula) <= 255:  # Excel inline-list formula1 hard limit
                dv = DataValidation(type="list", formula1=formula, allow_blank=not question.mandatory_flag)
                ws.add_data_validation(dv)
                dv.add(f"D{row}")
            # else: too many/long options for an inline dropdown; the
            # AllowedValues column still documents them as free text and
            # field validation below still enforces membership.

        for col in range(1, len(QUESTIONNAIRE_COLUMNS) + 1):
            ws.cell(row=row, column=col).protection = Protection(locked=col not in QUESTIONNAIRE_EDITABLE_COLUMNS)
        row += 1

    for col_letter in QUESTIONNAIRE_HIDDEN_COLUMNS:
        ws.column_dimensions[col_letter].hidden = True
    ws.protection.sheet = True


# ---------------------------------------------------------------------------
# Import engine (FS Section 16)
# ---------------------------------------------------------------------------


def parse_and_validate_workbook(
    file_bytes: bytes,
    expected_registration: "SupplierRegistration",
    *,
    expected_hash: str,
    expected_template_version: str,
    expected_questionnaire_version: str,
    expected_sheets: list[str],
    templates_by_module_code: dict[str, "TemplateDefinition"],
) -> ImportResult:
    """Validate a returned workbook per FS 16.2: structure first, then fields.

    A structural failure (wrong sheet set, headers, hidden columns, locked
    cell values, version, or hash) stops the pipeline before a single field
    is read -- a tampered file is never partially trusted. Only a
    structurally clean file proceeds to FS 15.2 field-level checks.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes))
    except Exception as exc:  # openpyxl raises several exception types for bad files
        failure = ValidationFailure(
            category="structural",
            sheet=None,
            cell=None,
            rule="unreadable_workbook",
            expected="a valid .xlsx file",
            actual=str(exc),
        )
        return _finalize_result([failure], {}, {})

    structural_failures = _validate_structure(
        wb,
        expected_registration=expected_registration,
        expected_hash=expected_hash,
        expected_template_version=expected_template_version,
        expected_questionnaire_version=expected_questionnaire_version,
        expected_sheets=expected_sheets,
        templates_by_module_code=templates_by_module_code,
    )
    if structural_failures:
        return _finalize_result(structural_failures, {}, {})

    supplier_info, info_failures = _extract_supplier_info(wb)
    answers_by_module, field_failures = _extract_and_validate_answers(wb, templates_by_module_code)
    failures = info_failures + field_failures

    return _finalize_result(failures, supplier_info, answers_by_module)


def _finalize_result(
    failures: list[ValidationFailure],
    supplier_info: dict[str, Any],
    answers_by_module: dict[str, dict[str, Any]],
) -> ImportResult:
    ok = not failures
    error_report_bytes = None if ok else build_error_report_xlsx(failures)
    summary = build_import_summary(failures, supplier_info=supplier_info, answers_by_module=answers_by_module)
    return ImportResult(
        ok=ok,
        failures=failures,
        supplier_info=supplier_info,
        answers_by_module=answers_by_module,
        error_report_bytes=error_report_bytes,
        import_summary_text=summary,
    )


def _validate_structure(
    wb: Workbook,
    *,
    expected_registration: "SupplierRegistration",
    expected_hash: str,
    expected_template_version: str,
    expected_questionnaire_version: str,
    expected_sheets: list[str],
    templates_by_module_code: dict[str, "TemplateDefinition"],
) -> list[ValidationFailure]:
    failures: list[ValidationFailure] = []

    actual_sheets = list(wb.sheetnames)
    if set(actual_sheets) != set(expected_sheets):
        failures.append(
            ValidationFailure(
                category="structural",
                sheet=None,
                cell=None,
                rule="sheet_names",
                expected=", ".join(sorted(expected_sheets)),
                actual=", ".join(sorted(actual_sheets)),
            )
        )
        # Can't safely inspect headers/cells on an unexpected sheet set.
        return failures

    if SHEET_SUPPLIER_INFO in actual_sheets:
        ws = wb[SHEET_SUPPLIER_INFO]
        headers = [ws.cell(row=1, column=c).value for c in range(1, len(SUPPLIER_INFO_COLUMNS) + 1)]
        if headers != list(SUPPLIER_INFO_COLUMNS):
            failures.append(
                ValidationFailure(
                    category="structural",
                    sheet=SHEET_SUPPLIER_INFO,
                    cell="1:1",
                    rule="column_headers",
                    expected=",".join(SUPPLIER_INFO_COLUMNS),
                    actual=",".join(str(h) for h in headers),
                )
            )

        actual_supplier_id = ws["A2"].value
        expected_supplier_id = str(expected_registration.id)
        if str(actual_supplier_id) != expected_supplier_id:
            failures.append(
                ValidationFailure(
                    category="tampered_locked_cell",
                    sheet=SHEET_SUPPLIER_INFO,
                    cell="A2",
                    rule="SupplierID",
                    expected=expected_supplier_id,
                    actual=str(actual_supplier_id),
                )
            )

        actual_template_version = ws["K2"].value
        if actual_template_version != expected_template_version:
            failures.append(
                ValidationFailure(
                    category="version_mismatch",
                    sheet=SHEET_SUPPLIER_INFO,
                    cell="K2",
                    rule="TemplateVersion",
                    expected=expected_template_version,
                    actual=str(actual_template_version),
                )
            )
        # SupplierType (B2) is intentionally NOT compared against a
        # ground-truth value here: this function's signature has no
        # supplier_type_code to compare against (SupplierRegistration has no
        # loaded relationship for it). It is still covered transitively --
        # any edit to B2 changes the structure hash checked below.

    if SHEET_INSTRUCTIONS in actual_sheets:
        ws = wb[SHEET_INSTRUCTIONS]
        actual_qv = ws["B3"].value
        if actual_qv != expected_questionnaire_version:
            failures.append(
                ValidationFailure(
                    category="version_mismatch",
                    sheet=SHEET_INSTRUCTIONS,
                    cell="B3",
                    rule="QuestionnaireVersion",
                    expected=expected_questionnaire_version,
                    actual=str(actual_qv),
                )
            )

    for module_code, template in templates_by_module_code.items():
        sheet_name = _module_sheet_name(module_code)
        if sheet_name not in actual_sheets:
            continue  # already reported by the sheet_names check above
        ws = wb[sheet_name]

        headers = [ws.cell(row=1, column=c).value for c in range(1, len(QUESTIONNAIRE_COLUMNS) + 1)]
        if headers != list(QUESTIONNAIRE_COLUMNS):
            failures.append(
                ValidationFailure(
                    category="structural",
                    sheet=sheet_name,
                    cell="1:1",
                    rule="column_headers",
                    expected=",".join(QUESTIONNAIRE_COLUMNS),
                    actual=",".join(str(h) for h in headers),
                )
            )

        for col_letter in QUESTIONNAIRE_HIDDEN_COLUMNS:
            if not ws.column_dimensions[col_letter].hidden:
                failures.append(
                    ValidationFailure(
                        category="structural",
                        sheet=sheet_name,
                        cell=f"{col_letter}:{col_letter}",
                        rule="hidden_column_visible",
                        expected="hidden",
                        actual="visible",
                    )
                )

        known_keys = {q.question_key for section in template.sections for q in section.questions}
        row = 2
        while ws.cell(row=row, column=1).value not in (None, ""):
            question_id = ws.cell(row=row, column=1).value
            module_id = ws.cell(row=row, column=2).value
            if module_id != module_code:
                failures.append(
                    ValidationFailure(
                        category="unknown_module",
                        sheet=sheet_name,
                        cell=f"B{row}",
                        rule="ModuleID",
                        expected=module_code,
                        actual=str(module_id),
                    )
                )
            if question_id not in known_keys:
                failures.append(
                    ValidationFailure(
                        category="unknown_question",
                        sheet=sheet_name,
                        cell=f"A{row}",
                        rule="QuestionID",
                        expected="one of " + ",".join(sorted(known_keys)),
                        actual=str(question_id),
                    )
                )
            score_formula = ws.cell(row=row, column=7).value
            if score_formula != SCORE_FORMULA_PLACEHOLDER:
                failures.append(
                    ValidationFailure(
                        category="tampered_locked_cell",
                        sheet=sheet_name,
                        cell=f"G{row}",
                        rule="ScoreFormula",
                        expected=SCORE_FORMULA_PLACEHOLDER,
                        actual=str(score_formula),
                    )
                )
            row += 1

    sheet_meta = _extract_sheet_meta(wb)
    actual_hash = compute_structure_hash(canonical_structure_payload(sheet_meta))
    if actual_hash != expected_hash:
        failures.append(
            ValidationFailure(
                category="tampered_locked_cell",
                sheet=None,
                cell=None,
                rule="structure_hash",
                expected=expected_hash,
                actual=actual_hash,
            )
        )

    return failures


def _extract_supplier_info(wb: Workbook) -> tuple[dict[str, Any], list[ValidationFailure]]:
    failures: list[ValidationFailure] = []
    ws = wb[SHEET_SUPPLIER_INFO]
    info: dict[str, Any] = {}
    for idx, header in enumerate(SUPPLIER_INFO_COLUMNS, start=1):
        info[header] = ws.cell(row=2, column=idx).value

    for header in SUPPLIER_INFO_MANDATORY:
        value = info.get(header)
        if value is None or str(value).strip() == "":
            failures.append(
                ValidationFailure(
                    category="mandatory_missing",
                    sheet=SHEET_SUPPLIER_INFO,
                    cell=None,
                    rule=header,
                    expected="non-empty",
                    actual="blank",
                )
            )

    email = info.get("ContactEmail")
    if email and not EMAIL_REGEX.match(str(email)):
        failures.append(
            ValidationFailure(
                category="invalid_format",
                sheet=SHEET_SUPPLIER_INFO,
                cell=None,
                rule="ContactEmail",
                expected="valid email address",
                actual=str(email),
            )
        )

    country = info.get("Country")
    if country and str(country).upper() not in ISO_COUNTRY_CODES:
        failures.append(
            ValidationFailure(
                category="invalid_format",
                sheet=SHEET_SUPPLIER_INFO,
                cell=None,
                rule="Country",
                expected="ISO 3166-1 alpha-2 code",
                actual=str(country),
            )
        )

    account = info.get("BankAccountNumber")
    if account and not BANK_ACCOUNT_REGEX.match(str(account)):
        failures.append(
            ValidationFailure(
                category="invalid_format",
                sheet=SHEET_SUPPLIER_INFO,
                cell=None,
                rule="BankAccountNumber",
                expected=BANK_ACCOUNT_REGEX.pattern,
                actual=str(account),
            )
        )

    routing = info.get("BankRoutingNumber")
    if routing and not BANK_ROUTING_REGEX.match(str(routing)):
        failures.append(
            ValidationFailure(
                category="invalid_format",
                sheet=SHEET_SUPPLIER_INFO,
                cell=None,
                rule="BankRoutingNumber",
                expected=BANK_ROUTING_REGEX.pattern,
                actual=str(routing),
            )
        )

    return info, failures


def _extract_and_validate_answers(
    wb: Workbook, templates_by_module_code: dict[str, "TemplateDefinition"]
) -> tuple[dict[str, dict[str, Any]], list[ValidationFailure]]:
    """Read Response/Comments per module and run FS 15.2 field checks.

    Only reached once _validate_structure() has passed with zero failures
    (including the unknown_question/unknown_module checks), so every
    question_key encountered here is guaranteed to resolve in
    templates_by_module_code -- no defensive None-handling needed for that.
    """
    failures: list[ValidationFailure] = []
    answers_by_module: dict[str, dict[str, Any]] = {}

    for module_code, template in templates_by_module_code.items():
        sheet_name = _module_sheet_name(module_code)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        questions_by_key = {q.question_key: q for section in template.sections for q in section.questions}
        answers: dict[str, Any] = {}

        row = 2
        while ws.cell(row=row, column=1).value not in (None, ""):
            question_key = ws.cell(row=row, column=1).value
            question = questions_by_key[question_key]
            response = ws.cell(row=row, column=4).value
            mandatory = str(ws.cell(row=row, column=6).value or "").strip().lower() == "yes"
            is_blank = response is None or str(response).strip() == ""

            if mandatory and is_blank:
                failures.append(
                    ValidationFailure(
                        category="mandatory_missing",
                        sheet=sheet_name,
                        cell=f"D{row}",
                        rule=str(question_key),
                        expected="non-empty",
                        actual="blank",
                    )
                )

            if not is_blank:
                if question.question_type == "numeric":
                    try:
                        float(str(response))
                    except (TypeError, ValueError):
                        failures.append(
                            ValidationFailure(
                                category="invalid_format",
                                sheet=sheet_name,
                                cell=f"D{row}",
                                rule=str(question_key),
                                expected="numeric value",
                                actual=str(response),
                            )
                        )

                options = question.options or (["Yes", "No"] if question.question_type == "yes_no" else None)
                if options and str(response) not in [str(o) for o in options]:
                    failures.append(
                        ValidationFailure(
                            category="invalid_dropdown",
                            sheet=sheet_name,
                            cell=f"D{row}",
                            rule=str(question_key),
                            expected=", ".join(str(o) for o in options),
                            actual=str(response),
                        )
                    )
                answers[str(question_key)] = response
            row += 1

        answers_by_module[module_code] = answers

    return answers_by_module, failures


# ---------------------------------------------------------------------------
# Error / summary artifacts (FS Section 15.4 / 16.5)
# ---------------------------------------------------------------------------


def build_error_report_xlsx(failures: list[ValidationFailure]) -> bytes:
    """ErrorReport.xlsx: one row per validation failure."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    headers = ["Category", "Sheet", "Cell", "Rule", "Expected", "Actual"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
    for row_idx, failure in enumerate(failures, start=2):
        ws.cell(row=row_idx, column=1, value=failure.category)
        ws.cell(row=row_idx, column=2, value=failure.sheet or "")
        ws.cell(row=row_idx, column=3, value=failure.cell or "")
        ws.cell(row=row_idx, column=4, value=failure.rule)
        ws.cell(row=row_idx, column=5, value=failure.expected or "")
        ws.cell(row=row_idx, column=6, value=failure.actual or "")
    for col_letter in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col_letter].width = 28

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_import_summary(
    failures: list[ValidationFailure],
    *,
    supplier_info: Optional[dict[str, Any]] = None,
    answers_by_module: Optional[dict[str, dict[str, Any]]] = None,
) -> str:
    """ImportSummary.txt: human-readable pass/fail summary for the SLP Admin."""
    timestamp = datetime.now(timezone.utc).isoformat()
    lines = ["S2PNexus Supplier Registration Import Summary", f"Generated: {timestamp}", ""]

    if not failures:
        module_count = len(answers_by_module or {})
        answer_count = sum(len(a) for a in (answers_by_module or {}).values())
        lines.append("Result: SUCCESS")
        lines.append(f"Modules imported: {module_count}")
        lines.append(f"Responses captured: {answer_count}")
        if supplier_info:
            lines.append(f"Supplier: {supplier_info.get('LegalName', '')}")
        return "\n".join(lines)

    by_category: dict[str, int] = {}
    for failure in failures:
        by_category[failure.category] = by_category.get(failure.category, 0) + 1
    lines.append("Result: FAILED")
    lines.append(f"Total errors: {len(failures)}")
    for category, count in sorted(by_category.items()):
        lines.append(f"  {category}: {count}")
    lines.append("")
    lines.append("See ErrorReport.xlsx for the full row-by-row detail.")
    return "\n".join(lines)
